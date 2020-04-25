import os
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

def menu_carpetas(fname, mensaje):
    print(mensaje)
    command = str(input('''¿Qué quieres hacer?: 
                        
                        [c]ontinuar, ya tengo el archivo.
                        [e]ditar el nombre del archivo, está mal.
                        [s]alir, aún no tengo el archivo.
                        
                        >>Ingresa una opción: '''))

    if command == 'c':
        existe_archivo(fname)
    elif command == 'e':
        nombre_archivo()
    else:
        print("Se salió del programa. No se crearon las carpetas. No se tiene el archivo aún.")
        exit()

def existe_archivo(fname):
    if os.path.isfile(fname):
      
      try:
        wb = load_workbook(filename=fname, read_only=True)
      except InvalidFileException:
        menu_carpetas(fname, mensaje="El archivo no tiene extensión .xlsx, se requiere un archivo con esta extensión.")
      else:
        crear_carpetas(wb)
        #informe_proceso
      finally:
        pass

    else:
      menu_carpetas(fname, mensaje="No se encontró el archivo '" + fname + "'. Se requiere para continuar.")

def nombre_archivo():
    fname = str(input("Escriba el nombre del archivo que contiene la lista de carpetas a crear: "))
    existe_archivo(fname)

def crear_carpetas(wb):
    grupos = wb.sheetnames
    for hoja in grupos:
      cantidad = 0
      grupo = wb[hoja]
      carpeta = grupo.title

      try:
        os.stat(carpeta)
      except:
        os.mkdir(carpeta)

      for fila in grupo.rows:
        for columna in fila:
          nombre = columna.value
          nombre = nombre.strip()
          subcarpeta = carpeta + os.sep + nombre

          try:
            os.stat(subcarpeta)
          except:
            os.mkdir(subcarpeta)
          
          cantidad = cantidad + 1
    
    print("")
    print("Se crearon {} carpetas.".format(cantidad))

if __name__ == "__main__":
    print("Programa para crear carpetas a partir de una lista en un archivo de Excel con extensión .xlsx")
    nombre_archivo()